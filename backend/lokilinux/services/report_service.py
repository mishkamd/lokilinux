"""
LokiLinux — Compliance Reporting Engine (docs/compliance/13-OPS.md Phase 5,
05-API.md §7, 01-DATA-MODEL.md §8).

Report data is computed on demand from `rule_evaluations` (confirmed
populated by lokilinux-compliance's Ingester.evaluateRules) rather than
`compliance_scores` (declared in migration 016 but never written by
anything, the same "table exists, no writer" gap this module has hit
before for drift_events/file_hashes) — reading the real evaluation table
directly avoids depending on a second table nothing keeps in sync, and
matches this module's "coverage is real, never silently assumed" ethos.

Category mapping matches docs/compliance/07-POLICY-ENGINE.md §4 exactly,
minus `packages` — that category sources from the existing `packages`
table (package inventory), a different subsystem this report generator
doesn't reach into yet; `overall` here is the mean of the four rule-backed
categories, not all five from the full spec.
"""

import csv
import io
import json
from datetime import datetime, timezone
from uuid import UUID

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from lokilinux.config import get_settings
from lokilinux.models.agent import Agent
from lokilinux.models.compliance_exception import ComplianceException
from lokilinux.models.compliance_report import ComplianceReport
from lokilinux.models.compliance_rule import ComplianceRule, PolicySetRule
from lokilinux.models.rule_evaluation import RuleEvaluation
from lokilinux.object_storage import ObjectStorage
from lokilinux.services.storage_service import StorageService
from lokilinux.settings_schema import get_setting_value

CATEGORY_BY_DOMAIN = {
    "sshd": "security",
    "pam": "security",
    "auditd": "security",
    "sudo": "security",
    "selinux": "security",
    "firewall": "security",
    "sysctl": "configuration",
    "systemd_services": "configuration",
    "cron": "configuration",
    "login_defs": "configuration",
    "password_policy": "configuration",
    "network": "configuration",
    "time_sync": "configuration",
    "mounts": "filesystem",
    "file_integrity": "filesystem",
    "repositories": "filesystem",
    "kernel": "kernel",
    "kernel_modules": "kernel",
}


async def _latest_evaluations(
    db: AsyncSession, agent_id: UUID | None = None, policy_set_id: UUID | None = None
):
    """One row per (agent_id, rule_id) — the most recent verdict, since
    rule_evaluations is an append-only hypertable that can carry many
    historical rows for the same rule. policy_set_id restricts to only the
    rules that belong to that one policy set (via policy_set_rules) — the
    join a POLICY_SET-scoped report needs instead of fleet-wide data.
    """
    q = (
        select(RuleEvaluation, ComplianceRule.domain, ComplianceRule.severity, ComplianceRule.title, Agent.hostname)
        .join(ComplianceRule, ComplianceRule.id == RuleEvaluation.rule_id)
        .outerjoin(Agent, Agent.id == RuleEvaluation.agent_id)
        .distinct(RuleEvaluation.agent_id, RuleEvaluation.rule_id)
        .order_by(RuleEvaluation.agent_id, RuleEvaluation.rule_id, RuleEvaluation.time.desc())
    )
    if agent_id:
        q = q.where(RuleEvaluation.agent_id == agent_id)
    if policy_set_id:
        q = q.join(PolicySetRule, PolicySetRule.rule_id == ComplianceRule.id).where(
            PolicySetRule.policy_set_id == policy_set_id
        )
    return (await db.execute(q)).all()


async def build_fleet_summary_data(
    db: AsyncSession, agent_id: UUID | None = None, policy_set_id: UUID | None = None
) -> dict:
    rows = await _latest_evaluations(db, agent_id, policy_set_id)

    category_counts: dict[str, dict[str, int]] = {}
    violations: list[dict] = []

    for evaluation, domain, severity, title, hostname in rows:
        category = CATEGORY_BY_DOMAIN.get(domain, "configuration")
        bucket = category_counts.setdefault(category, {"passed": 0, "failed": 0})
        if evaluation.result == "PASS":
            bucket["passed"] += 1
        elif evaluation.result == "FAIL":
            bucket["failed"] += 1
            violations.append(
                {
                    "agent_id": str(evaluation.agent_id),
                    "hostname": hostname,
                    "domain": domain,
                    "severity": severity,
                    "title": title,
                    "time": evaluation.time.isoformat(),
                }
            )
        # ERROR/NOT_APPLICABLE/NOT_EVALUATED excluded from both numerator and
        # denominator — same rule 08-POLICY-ENGINE.md §4 applies fleet-wide.

    categories = {}
    for category, counts in category_counts.items():
        total = counts["passed"] + counts["failed"]
        categories[category] = {
            **counts,
            "score": round(100.0 * counts["passed"] / total, 1) if total else None,
        }

    scored = [c["score"] for c in categories.values() if c["score"] is not None]
    overall_score = round(sum(scored) / len(scored), 1) if scored else None

    violations.sort(
        key=lambda v: {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}.get(v["severity"], 4)
    )

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "overall_score": overall_score,
        "categories": categories,
        "top_violations": violations[:50],
        "total_rules_evaluated": len(rows),
    }


async def build_framework_report_data(
    db: AsyncSession, framework_key: str, policy_set_id: UUID | None = None
) -> dict:
    """Control-level pass/fail across the fleet for one framework
    (docs/compliance §30, §19) — groups the same latest-verdict-per-rule
    data build_fleet_summary_data reads, but by Control instead of domain
    category, via the normalized Framework/Control/RuleMapping tables (F9)."""
    rows = await _latest_evaluations(db, policy_set_id=policy_set_id)

    control_rows = (
        await db.execute(
            text(
                """
                SELECT rm.rule_id, c.control_id, c.title
                FROM compliance_rule_mappings rm
                JOIN compliance_controls c ON c.id = rm.control_id
                JOIN compliance_framework_versions fv ON fv.id = c.framework_version_id
                JOIN compliance_frameworks f ON f.id = fv.framework_id
                WHERE f.key = :key
                """
            ),
            {"key": framework_key},
        )
    ).mappings().all()
    rule_to_controls: dict[UUID, list[dict]] = {}
    for r in control_rows:
        rule_to_controls.setdefault(r["rule_id"], []).append({"control_id": r["control_id"], "title": r["title"]})

    control_counts: dict[str, dict] = {}
    for evaluation, _domain, _severity, _title, _hostname in rows:
        for c in rule_to_controls.get(evaluation.rule_id, []):
            bucket = control_counts.setdefault(c["control_id"], {"title": c["title"], "passed": 0, "failed": 0})
            if evaluation.result == "PASS":
                bucket["passed"] += 1
            elif evaluation.result == "FAIL":
                bucket["failed"] += 1

    controls = []
    for control_id, b in sorted(control_counts.items()):
        total = b["passed"] + b["failed"]
        controls.append(
            {
                "control_id": control_id,
                "title": b["title"],
                "passed": b["passed"],
                "failed": b["failed"],
                "score": round(100.0 * b["passed"] / total, 1) if total else None,
            }
        )
    scored = [c["score"] for c in controls if c["score"] is not None]

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "framework": framework_key,
        "overall_score": round(sum(scored) / len(scored), 1) if scored else None,
        "controls": controls,
    }


async def build_exception_report_data(db: AsyncSession) -> dict:
    """Every compliance exception, any status — docs/compliance §17:
    "Exceptions must appear in reports and audit logs.\""""
    rows = (await db.execute(select(ComplianceException))).scalars().all()
    exceptions = [
        {
            "id": str(e.id),
            "rule_id": str(e.rule_id),
            "agent_id": str(e.agent_id) if e.agent_id else None,
            "reason": e.reason,
            "owner": e.owner,
            "status": e.status,
            "expires_at": e.expires_at.isoformat(),
            "approved_by": str(e.approved_by) if e.approved_by else None,
        }
        for e in rows
    ]
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total": len(exceptions),
        "active": sum(1 for e in exceptions if e["status"] == "ACTIVE"),
        "exceptions": exceptions,
    }


async def build_executive_summary_data(db: AsyncSession) -> dict:
    """Leadership-facing rollup (docs/compliance §31) built entirely from
    already-stored evidence — no AI text generation here (§30: "AI must
    never invent findings"); this is the real structured summary a future
    AI-narrative layer would read from, not a substitute for one."""
    fleet = await build_fleet_summary_data(db)
    exceptions = await build_exception_report_data(db)
    servers_evaluated = (
        await db.execute(select(func.count(func.distinct(RuleEvaluation.agent_id))))
    ).scalar_one()

    return {
        "generated_at": fleet["generated_at"],
        "servers_evaluated": servers_evaluated,
        "overall_score": fleet["overall_score"],
        "categories": fleet["categories"],
        "top_violations": fleet["top_violations"][:10],
        "exceptions_active": exceptions["active"],
        "exceptions_total": exceptions["total"],
    }


def _flatten_for_csv(data: dict) -> list[dict]:
    rows = [
        {
            "row_type": "category",
            "name": "overall",
            "score": data["overall_score"],
            "passed": "",
            "failed": "",
        }
    ]
    for name, c in data["categories"].items():
        rows.append(
            {
                "row_type": "category",
                "name": name,
                "score": c["score"],
                "passed": c["passed"],
                "failed": c["failed"],
            }
        )
    for v in data["top_violations"]:
        rows.append(
            {
                "row_type": "violation",
                "name": v["title"],
                "score": v["severity"],
                "passed": v["domain"],
                "failed": v["agent_id"],
            }
        )
    return rows


def to_json(data: dict) -> bytes:
    return json.dumps(data, indent=2).encode("utf-8")


def to_csv(data: dict) -> bytes:
    buf = io.StringIO()
    rows = _flatten_for_csv(data)
    writer = csv.DictWriter(buf, fieldnames=["row_type", "name", "score", "passed", "failed"])
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def to_xlsx(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.worksheets[
        0
    ]  # Workbook() always creates exactly one sheet — index, not .active, avoids an Optional type
    ws.title = "Compliance Report"
    ws.append(["Category", "Score", "Passed", "Failed"])
    ws.append(["overall", data["overall_score"], "", ""])
    for name, c in data["categories"].items():
        ws.append([name, c["score"], c["passed"], c["failed"]])

    ws2 = wb.create_sheet("Top Violations")
    ws2.append(["Severity", "Domain", "Title", "Agent", "Time"])
    for v in data["top_violations"]:
        ws2.append([v["severity"], v["domain"], v["title"], v["agent_id"], v["time"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_TABLE_STYLE = TableStyle(
    [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]
)


def to_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    overall = data["overall_score"] if data["overall_score"] is not None else "N/A"
    story = [
        Paragraph("Compliance Report", styles["Title"]),
        Paragraph(f"Generated: {data['generated_at']}", styles["Normal"]),
        Paragraph(f"Overall score: {overall}", styles["Heading2"]),
        Spacer(1, 12),
    ]

    category_rows = [["Category", "Score", "Passed", "Failed"]]
    for name, c in data["categories"].items():
        category_rows.append(
            [name, c["score"] if c["score"] is not None else "N/A", c["passed"], c["failed"]]
        )
    category_table = Table(category_rows, hAlign="LEFT")
    category_table.setStyle(_TABLE_STYLE)
    story += [category_table, Spacer(1, 20), Paragraph("Top Violations", styles["Heading2"])]

    violations = data["top_violations"]
    if violations:
        violation_rows = [["Severity", "Domain", "Title", "Agent"]]
        violation_rows += [
            [v["severity"], v["domain"], v["title"], v["agent_id"]] for v in violations
        ]
        violation_table = Table(violation_rows, hAlign="LEFT")
        violation_table.setStyle(_TABLE_STYLE)
        story.append(violation_table)
    else:
        story.append(Paragraph("No violations.", styles["Normal"]))

    doc.build(story)
    return buf.getvalue()


def _generic_rows(data: dict) -> list[dict]:
    """Flattens an arbitrary report dict for CSV: each item of a top-level
    list-of-dicts field becomes one row tagged with its section name; every
    top-level scalar field becomes one "summary" row. Used by the FRAMEWORK/
    EXCEPTION/EXECUTIVE_SUMMARY report types, whose shapes vary too much for
    the fixed categories/top_violations columns _flatten_for_csv assumes."""
    rows = []
    for key, value in data.items():
        if isinstance(value, list) and value and isinstance(value[0], dict):
            for item in value:
                row = {"section": key}
                row.update({k: v for k, v in item.items() if not isinstance(v, (list, dict))})
                rows.append(row)
        elif not isinstance(value, (list, dict)):
            rows.append({"section": "summary", "field": key, "value": value})
    return rows


def to_csv_generic(data: dict) -> bytes:
    rows = _generic_rows(data)
    if not rows:
        return b""
    fieldnames = sorted({k for r in rows for k in r})
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def to_xlsx_generic(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "Summary"
    for key, value in data.items():
        if not isinstance(value, (list, dict)):
            ws.append([key, value])
    for key, value in data.items():
        if isinstance(value, list) and value and isinstance(value[0], dict):
            sheet = wb.create_sheet(key[:31])
            headers = list(value[0].keys())
            sheet.append(headers)
            for item in value:
                sheet.append([item.get(h, "") for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf_generic(data: dict, title: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story: list = [Paragraph(title, styles["Title"])]
    for key, value in data.items():
        if isinstance(value, list) and value and isinstance(value[0], dict):
            story.append(Spacer(1, 12))
            story.append(Paragraph(key.replace("_", " ").title(), styles["Heading2"]))
            headers = list(value[0].keys())
            table_rows = [headers] + [[str(item.get(h, "")) for h in headers] for item in value]
            table = Table(table_rows, hAlign="LEFT")
            table.setStyle(_TABLE_STYLE)
            story.append(table)
        elif not isinstance(value, (list, dict)):
            story.append(Paragraph(f"{key.replace('_', ' ').title()}: {value}", styles["Normal"]))
    doc.build(story)
    return buf.getvalue()


FORMAT_CONTENT_TYPES = {
    "JSON": "application/json",
    "CSV": "text/csv",
    "XLSX": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "PDF": "application/pdf",
}

_SERIALIZERS = {"JSON": to_json, "CSV": to_csv, "XLSX": to_xlsx, "PDF": to_pdf}
_GENERIC_REPORT_TYPES = {"FRAMEWORK", "EXCEPTION", "EXECUTIVE_SUMMARY"}


async def generate_report(
    db: AsyncSession, storage: ObjectStorage, report: ComplianceReport
) -> None:
    """Runs as a FastAPI BackgroundTask (see routers/compliance/reports.py)
    — mutates and commits `report` in place, matching the same
    Job-row-as-progress-tracker pattern the ComplianceAsCode importer uses.

    Writes the generated artifact to object storage (Object Storage plan)
    instead of the legacy `body` BYTEA column — see
    models/compliance_report.py's dual-read docstring.
    """
    report.status = "GENERATING"
    await db.commit()

    try:
        xlsx_pdf_enabled = await get_setting_value(db, "reports.xlsx_pdf_enabled")
        if report.format in ("XLSX", "PDF") and not xlsx_pdf_enabled:
            raise ValueError(
                "XLSX/PDF reports are disabled — enable reports.xlsx_pdf_enabled in Settings, "
                "or request JSON/CSV instead"
            )
        agent_id = UUID(report.params["agent_id"]) if report.params.get("agent_id") else None
        if report.report_type == "POLICY_SET":
            if not report.params.get("policy_set_id"):
                raise ValueError("POLICY_SET reports require params.policy_set_id")
            policy_set_id = UUID(report.params["policy_set_id"])
            data = await build_fleet_summary_data(
                db, agent_id=agent_id, policy_set_id=policy_set_id
            )
        elif report.report_type in ("FLEET_SUMMARY", "DATACENTER", "CUSTOM"):
            data = await build_fleet_summary_data(db, agent_id=agent_id)
        elif report.report_type == "FRAMEWORK":
            if not report.params.get("framework"):
                raise ValueError("FRAMEWORK reports require params.framework")
            policy_set_id = UUID(report.params["policy_set_id"]) if report.params.get("policy_set_id") else None
            data = await build_framework_report_data(db, report.params["framework"], policy_set_id)
        elif report.report_type == "EXCEPTION":
            data = await build_exception_report_data(db)
        elif report.report_type == "EXECUTIVE_SUMMARY":
            data = await build_executive_summary_data(db)
        else:
            raise ValueError(f"unknown report_type {report.report_type!r}")

        if report.report_type in _GENERIC_REPORT_TYPES:
            generic_serializers = {
                "JSON": to_json,
                "CSV": to_csv_generic,
                "XLSX": to_xlsx_generic,
                "PDF": lambda d: to_pdf_generic(d, report.report_type.replace("_", " ").title() + " Report"),
            }
            serialize = generic_serializers[report.format]
        else:
            serialize = _SERIALIZERS[report.format]
        body = serialize(data)

        ext = report.format.lower()
        obj = await StorageService(storage, db).store_bytes(
            body,
            category="compliance.report",
            original_filename=f"compliance-report-{report.id}.{ext}",
            content_type=FORMAT_CONTENT_TYPES[report.format],
            max_bytes=get_settings().s3_max_upload_bytes,
            created_by=report.generated_by,
        )
        report.storage_object_id = obj.id
        report.status = "COMPLETED"
        report.completed_at = datetime.now(timezone.utc)
        report.artifact_uri = f"/api/v1/compliance/reports/{report.id}/download"
    except Exception as exc:
        report.status = "FAILED"
        report.error_message = str(exc)
        report.completed_at = datetime.now(timezone.utc)

    await db.commit()
